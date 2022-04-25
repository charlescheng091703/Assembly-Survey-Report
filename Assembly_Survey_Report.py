import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from copy import copy
from typing import Union, Optional
import numpy as np
import xlwings as xw
import shutil
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
import os
import ipywidgets as widgets
from IPython.display import display, clear_output
from colorama import Fore, Style
from win32com import client
import win32api
import pathlib
from magnetModuleList import *
from datetime import date, datetime 

widget_out = widgets.Output(layout={'border': '1px solid black'})

########## Credit: MaxU on Stack Overflow ###########
##### https://www.linkedin.com/in/maxuzunov/ ########

def copy_excel_cell_range(
        src_ws: openpyxl.worksheet.worksheet.Worksheet,
        min_row: int = None,
        max_row: int = None,
        min_col: int = None,
        max_col: int = None,
        tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
        tgt_min_row: int = 1,
        tgt_min_col: int = 1,
        with_style: bool = True
) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    copies all cells from the source worksheet [src_ws] starting from [min_row] row
    and [min_col] column up to [max_row] row and [max_col] column
    to target worksheet [tgt_ws] starting from [tgt_min_row] row
    and [tgt_min_col] column.

    @param src_ws:  source worksheet
    @param min_row: smallest row index in the source worksheet (1-based index)
    @param max_row: largest row index in the source worksheet (1-based index)
    @param min_col: smallest column index in the source worksheet (1-based index)
    @param max_col: largest column index in the source worksheet (1-based index)
    @param tgt_ws:  target worksheet.
                    If None, then the copy will be done to the same (source) worksheet.
    @param tgt_min_row: target row index (1-based index)
    @param tgt_min_col: target column index (1-based index)
    @param with_style:  whether to copy cell style. Default: True

    @return: target worksheet object
    """
    if tgt_ws is None:
        tgt_ws = src_ws

    # https://stackoverflow.com/a/34838233/5741205
    for row in src_ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
        for cell in row:
            tgt_cell = tgt_ws.cell(
                row=cell.row + tgt_min_row - 1,
                column=cell.col_idx + tgt_min_col - 1,
                value=cell.value
            )
            if with_style and cell.has_style:
                # tgt_cell._style = copy(cell._style)
                tgt_cell.font = copy(cell.font)
                tgt_cell.border = copy(cell.border)
                tgt_cell.fill = copy(cell.fill)
                tgt_cell.number_format = copy(cell.number_format)
                tgt_cell.protection = copy(cell.protection)
                tgt_cell.alignment = copy(cell.alignment)
    return tgt_ws


def append_df_to_excel(
        filename: Union[str, Path],
        df: pd.DataFrame,
        sheet_name: str = 'Sheet1',
        startrow: Optional[int] = None,
        max_col_width: int = 30,
        autofilter: bool = False,
        fmt_int: str = "#,##0",
        fmt_float: str = "#,##0.00",
        fmt_date: str = "yyyy-mm-dd",
        fmt_datetime: str = "yyyy-mm-dd hh:mm",
        truncate_sheet: bool = False,
        storage_options: Optional[dict] = None,
        **to_excel_kwargs
) -> None:
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param max_col_width: maximum column width in Excel. Default: 40
    @param autofilter: boolean - whether add Excel autofilter or not. Default: False
    @param fmt_int: Excel format for integer numbers
    @param fmt_float: Excel format for float numbers
    @param fmt_date: Excel format for dates
    @param fmt_datetime: Excel format for datetime's
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param storage_options: dict, optional
        Extra options that make sense for a particular storage connection, e.g. host, port,
        username, password, etc., if using a URL that will be parsed by fsspec, e.g.,
        starting “s3://”, “gcs://”.
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('/tmp/test.xlsx', df, autofilter=True,
                           freeze_panes=(1,0))

    >>> append_df_to_excel('/tmp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('/tmp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    >>> append_df_to_excel('/tmp/test.xlsx', df, index=False,
                           fmt_datetime="dd.mm.yyyy hh:mm")

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    def set_column_format(ws, column_letter, fmt):
        for cell in ws[column_letter]:
            cell.number_format = fmt
    filename = Path(filename)
    file_exists = filename.is_file()
    # process parameters
    # calculate first column number
    # if the DF will be written using `index=True`, then `first_col = 2`, else `first_col = 1`
    first_col = int(to_excel_kwargs.get("index", True)) + 1
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    # save content of existing sheets
    if file_exists:
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        sheet_exists = sheet_name in sheet_names
        sheets = {ws.title: ws for ws in wb.worksheets}

    with pd.ExcelWriter(
        filename.with_suffix(".xlsx"),
        engine="openpyxl",
        mode="a" if file_exists else "w",
        if_sheet_exists="new" if file_exists else None,
        date_format=fmt_date,
        datetime_format=fmt_datetime,
        storage_options=storage_options
    ) as writer:
        if file_exists:
            # try to open an existing workbook
            writer.book = wb
            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
            # copy existing sheets
            writer.sheets = sheets
        else:
            # file doesn't exist, we are creating a new one
            startrow = 0

        # write out the DataFrame to an ExcelWriter
        df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)
        worksheet = writer.sheets[sheet_name]

        if autofilter:
            worksheet.auto_filter.ref = worksheet.dimensions

        for xl_col_no, dtyp in enumerate(df.dtypes, first_col):
            col_no = xl_col_no - first_col
            width = max(df.iloc[:, col_no].astype(str).str.len().max(),
                        len(df.columns[col_no]) + 6)
            width = min(max_col_width, width)
            column_letter = get_column_letter(xl_col_no)
            worksheet.column_dimensions[column_letter].width = width
            if np.issubdtype(dtyp, np.integer):
                set_column_format(worksheet, column_letter, fmt_int)
            if np.issubdtype(dtyp, np.floating):
                set_column_format(worksheet, column_letter, fmt_float)

    if file_exists and sheet_exists:
        # move (append) rows from new worksheet to the `sheet_name` worksheet
        wb = load_workbook(filename)
        # retrieve generated worksheet name
        new_sheet_name = set(wb.sheetnames) - set(sheet_names)
        if new_sheet_name:
            new_sheet_name = list(new_sheet_name)[0]
        # copy rows written by `df.to_excel(...)` to
        copy_excel_cell_range(
            src_ws=wb[new_sheet_name],
            tgt_ws=wb[sheet_name],
            tgt_min_row=startrow + 1,
            with_style=True
        )
        # remove new (generated by Pandas) worksheet
        del wb[new_sheet_name]
        wb.save(filename)
        wb.close()

#####################################################
#####################################################

def read_csv(filename, col_names = False):
    
    if col_names:
        df = pd.read_csv(filename, index_col=0)
    else:
        df = pd.read_csv(filename, names=['Key','Value'], index_col=0)
    
    return df

def extract_csv_data(dataframe, index_names, column_names=['Value']):
    
    values = [[dataframe.loc[index][column] for column in column_names] for index in index_names]
    
    return values

def write_excel_col(filename, sheet_name, values, start_index='A1'):
    
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    
    col = start_index[0]
    row = int(start_index[1:])
    for i in range(len(values)):
        if type(values[i])== tuple:
            ws[col+str(row+i)].hyperlink = values[i][1]
            ws[col+str(row+i)].value = values[i][0]
            ws[col+str(row+i)].font = Font(name="Calibri", size=11, color="000645AD", underline="single")
            ws[col+str(row+i)].alignment = Alignment(horizontal='left')
            
        else:
            ws[col+str(row+i)] = values[i]
            ws[col+str(row+i)].font = Font(name="Calibri", size=11, color="00000000")
            ws[col+str(row+i)].alignment = Alignment(horizontal='left')
            
    wb.save(filename)
    
def write_excel_row(filename, sheet_name, values, start_index='A1'):
    
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    
    col = start_index[0]
    row = int(start_index[1:])
    for i in range(len(values)):
        ws[chr(ord(col)+i)+str(row)] = values[i]
        ws[chr(ord(col)+i)+str(row)].font = Font(name="Calibri", size=11, color="00000000")
        ws[chr(ord(col)+i)+str(row)].alignment = Alignment(horizontal='left')
            
    wb.save(filename)

def write_excel(filename, sheet_name, values, start_index):
    
    for i in range(len(values)):
        write_excel_col(filename, sheet_name, values[i], start_index[i])

def extract_excel_data(filename, columns, sheet_name=None, start_row=1, end_row=None,copy_formula=True):
    
    wb = load_workbook(filename, data_only=not copy_formula)
    if sheet_name==None:
        ws = wb.active
    else:
        ws = wb[sheet_name]
    if end_row==None:
        end_row = ws.max_row
        
    values = [[(ws.cell(i,col).value, ws.cell(i,col).hyperlink) if ws.cell(i,col).hyperlink!=None else ws.cell(i,col).value for i in range(start_row,end_row+1)] for col in columns]
    
    return values

def copy_paste_wrksht(workbook1, workbook2, sheet_name):
    
    app = xw.App(visible=False)
    wb1 = xw.Book(workbook1)
    wb2 = xw.Book(workbook2)

    ws1 = wb1.sheets(1)
    new_ws = ws1.copy(after=wb2.sheets[sheet_name])
    wb2.sheets[sheet_name].delete()
    new_ws.name = sheet_name
    wb2.save()
    app.kill()

def delete_images(workbook, sheet_name):
    
    wb = load_workbook(workbook)
    for sheet in sheet_name:
        wb[sheet]._images.clear()
    wb.save(workbook)

def stylize_cells(workbook, sheet_name, cell_bounds, align=None, number_decimals=False, backgrd_color=None, thick_right=None, thick_left=None, thick_top=None, thick_bottom=None, bold=False, num_indent=False, unbold=False, border=None):
    
    wb = load_workbook(workbook)
    ws = wb[sheet_name]
    thick = Side(border_style="thick", color="00000000")
    if border != None:
        thin = Side(border_style="thin", color="00D3D3D3")
    else:
        thin = Side(border_style="thin", color="00000000")
    
    border_format_exec = "ws.cell(row, col).border = Border("
    if thick_top != None:
        border_format_exec += "top=thick,"
    else:
        border_format_exec += "top=thin,"
    if thick_left != None:
        border_format_exec += "left=thick,"
    else:
        border_format_exec += "left=thin,"
    if thick_right != None:
        border_format_exec += "right=thick,"
    else:
        border_format_exec += "right=thin,"
    if thick_bottom != None:
        border_format_exec += "bottom=thick)"
    else:
        border_format_exec += "bottom=thin)"

    if number_decimals is not False:
        number_format_exec = "ws.cell(row, col).number_format = '0."
        for i in range(number_decimals):
            number_format_exec += "0"
        number_format_exec += "'"

    for col in range(ord(cell_bounds[0][0])-64,ord(cell_bounds[1][0])-63):
        for row in range(int(cell_bounds[0][1:]),int(cell_bounds[1][1:])+1):
            if align != None:
                ws.cell(row, col).alignment = Alignment(horizontal=align)
            if bold:
                ws.cell(row, col).font = Font(size=16,bold=True)
            if unbold:
                ws.cell(row, col).font = Font(size=11,bold=False)
            exec(border_format_exec)
            if number_decimals is not False:
                exec(number_format_exec)
            if num_indent is not False:
                ws.cell(row, col).alignment = Alignment(horizontal=align, indent=num_indent)
            if backgrd_color != None:
                ws.cell(row, col).fill = PatternFill(start_color=backgrd_color, end_color=backgrd_color, fill_type = "solid")
    wb.save(workbook)

def remove_rows(workbook, sheet_name, row_bounds='1:1'):
    
    app = xw.App()
    wb = app.books.open(workbook)
    wb.sheets[sheet_name].range(row_bounds).delete() 
    wb.save()
    app.kill()

def autosize_row_height(workbook, sheet_name,size=False):
    
    wb = load_workbook(workbook)
    ws = wb[sheet_name]
    
    rowHeights = [ws.row_dimensions[i+1].height for i in range(ws.max_row)]
    rowHeights = [15 if rh is None else rh for rh in rowHeights]
    
    if size is not False:
        row_height = 16
    else:
        row_height = 45
    
    for i, height in enumerate(rowHeights):
        if height > row_height:
            ws.row_dimensions[i+1].height = row_height
            
    wb.save(workbook)

def autofit_columns(workbook, sheet_name):
    
    wb = load_workbook(workbook)
    worksheet = wb[sheet_name]
    
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        if column != 'A':
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
        
    wb.save(workbook)

def no_fill(workbook, sheet_name):
    
    wb = load_workbook(workbook)
    ws = wb[sheet_name]
    
    no_fill = openpyxl.styles.PatternFill(fill_type=None)
    for row in ws:
        for cell in row:
            cell.fill = no_fill

    wb.save(workbook)

def log_entry(filename, data):
    
    wb = load_workbook("Report_Log.xlsx")
    ws = wb.active
    
    for row in range(1,ws.max_row+1):
        if ws.cell(row, 1).value == None:
            row_write = row
            break
            
    ws.cell(row_write, 1).value = filename
    ws.cell(row_write, 2).value = datetime.time(datetime.now())
    ws.cell(row_write, 3).value = date.today().strftime("%B %d, %Y")
    ws.cell(row_write, 4).value = data[0]
    ws.cell(row_write, 5).value = data[1]
    ws.cell(row_write, 6).value = data[2]

    ws.cell(row_write, 4).number_format = '0.000000'
    ws.cell(row_write, 5).number_format = '0.000000'
    ws.cell(row_write, 6).number_format = '0.000000'
    
    wb.save("Report_Log.xlsx")

def extract_RMS(workbook, sheet_name, bounds):

    app = xw.App()
    wb = app.books.open(workbook)
    data = wb.sheets[sheet_name].range(bounds).value
    wb.save()
    app.kill()

    return data

def extract_magnet_list(module_name):

    magnetlist_dict = read_data_test(module_name) 
    ordered = [(i,value["order"]) for i, value in enumerate(list(magnetlist_dict.values()))]
    ordered.sort(key = lambda x:x[1]) 
    ordered = [item[0] for item in ordered]
    magnetlist = list(magnetlist_dict)
    
    magnet_indicator = ['Q','F','M','S']
    magnet_names = {'Q':'Quadrupole','M':'Dipole Magnet','S':'Sextupole Magnet'}
    serial = []
    label = []
    url = []
    for index in ordered:
        item = magnetlist[index]
        try:
            if any([item[item.index(':')+1] == letter for letter in magnet_indicator]) and item.count(':') == 1:
                serial.append(magnetlist_dict[item]["serial"])
                label.append(magnetlist_dict[item]["label"])
                url.append((magnetlist_dict[item]["name"],magnetlist_dict[item]["url"]))
        except:
            pass
    
    return label, url, serial

def savefile_to_pdf(excel_file):

    pdf_file = excel_file[:-5] + '.pdf'
    excel_path = str(pathlib.Path.cwd() / excel_file)
    pdf_path = str(pathlib.Path.cwd() / pdf_file)

    excel = client.DispatchEx("Excel.Application")
    excel.Visible = 0

    wb = excel.Workbooks.Open(excel_path)
    wb.Worksheets([1]).Select()

    try:
        wb.ActiveSheet.ExportAsFixedFormat(0, pdf_path)
    except Exception as e:
        print("Failed to convert")
        print(str(e))
    finally:
        wb.Close()
        excel.Quit()
        
def cdb_to_survey(module_name):
    
    survey_prefixes = {'DLMA':'DA','DLMB':'DB','FODO':'FD','QMQA':'QA','QMQB':'QB'}
    module_type = module_name[0:4]
    module_number =  int(module_name[-4:])
    survey_number = int(module_number/10-100)
    if survey_number < 10:
        survey_number_str = '0' + str(survey_number)
    else:
        survey_number_str = str(survey_number)
    return survey_prefixes[module_type] + '_' + survey_number_str
        
@widget_out.capture()
def on_button_clicked(b):

    clear_output(wait=False)
    if len(module_name.value) == 0:
        print(Fore.RED + "Please enter the module name." + Style.RESET_ALL)
    else:
        generate_excel_report(module_name.value)

def generate_excel_report(module_name):
    
    print("\033[1mGenerating "+module_name+" Report...\033[0m")
    filename_report = module_name + '/Report ' + module_name + ' Assembly Survey.xlsx'
    filename_report = os.path.abspath(filename_report)
    module_type = module_name[0]
    if module_type == 'D':
        shutil.copy('Form_DLM_SurveyReport.xlsx', filename_report)
    elif module_type == 'F':
        shutil.copy('Form_FODO_SurveyReport.xlsx', filename_report)
    elif module_type == 'Q':
        shutil.copy('Form_QMQ_SurveyReport.xlsx', filename_report)
    
    try:
        df = read_csv(module_name+'/M_VERTEX.csv',col_names=True)
        M1_data = []
        M1_data.append(str(df.index.name))
        for i in df.columns:
            M1_data.append(float(i))
        write_excel_row(filename_report,'Alignment Summary',M1_data,'B41')
    except:
        print("M1 data excluded...")

    df = read_csv(module_name+'/INFO.csv')
    data = extract_csv_data(df,['Survey Date:','Surveyor(s):','Instrument s/n:','SA Version:','SA Filename:'])
    data[4][0] = data[4][0][data[4][0].rfind('\\')+1:]
    data = [item[0] for item in data]
    data.append(date.today().strftime("%B %d, %Y"))
    write_excel_col(filename_report,'Alignment Summary',data,'C3')
    write_excel_col(filename_report,'Alignment Summary',[module_name],'B1')
    print("Alignment Summary tab complete...")
    
    df = read_csv(module_name+'/CENTERS.csv',col_names=True)
    append_df_to_excel(filename_report,df,sheet_name="Alignment Summary",startcol=1,startrow=24)
    
    name, url, serial = extract_magnet_list(module_name)
    num_magnets = len(name)
    write_excel_col(filename_report, 'Alignment Summary', name, start_index='B11')
    write_excel_col(filename_report, 'Alignment Summary', url, start_index='C11')
    write_excel_col(filename_report, 'Alignment Summary', serial, start_index='E11')
    
    survey_name = cdb_to_survey(module_name)
    write_excel_col(filename_report,'Alignment Summary',[survey_name],'C1')
    
    if module_type == 'Q':
        RMS_data = ['N/A','N/A','N/A']
    else:
        RMS_data = extract_RMS(filename_report,'Alignment Summary','C36:E36')
    
    copy_paste_wrksht(module_name+'/FIDUCIALS.xls',filename_report,'Installation Fiducials')
    copy_paste_wrksht(module_name+'/TRANSFORMS.xls',filename_report,'Transformations')
    copy_paste_wrksht(module_name+'/USMN.xls',filename_report,'USMN Raw')
    print("Installation Fiducials tab complete...")
    print("Transformations tab complete...")
    print("USMN Raw tab complete...")
    
    delete_images(filename_report,['Installation Fiducials','Transformations','USMN Raw'])
    
    stylize_cells(filename_report,'Alignment Summary',['F26','H33'],align='right',number_decimals=3,num_indent=2)
    stylize_cells(filename_report,'Alignment Summary',['C26','E33'],align='center',number_decimals=6,backgrd_color='00ffffcd')
    stylize_cells(filename_report,'Alignment Summary',['B25','B33'],align='center',backgrd_color='00eef5e9')
    stylize_cells(filename_report,'Alignment Summary',['C25','H25'],align='center',backgrd_color='00eef5e9')
    stylize_cells(filename_report,'Alignment Summary',['H25','H33'])
    stylize_cells(filename_report,'Alignment Summary',['B1','B1'],bold=True,align='center',border=False)
    stylize_cells(filename_report,'Alignment Summary',['C1','C1'],bold=True,align='center',border=False)
    
    if module_type != 'F' and module_type != 'Q':
        stylize_cells(filename_report,'Alignment Summary',['B41','B41'],unbold=True,align='center',backgrd_color='00fedcd6',thick_left=True)
        stylize_cells(filename_report,'Alignment Summary',['C41','E41'],unbold=True,align='center',backgrd_color='00f2f2f2',number_decimals=6)
        stylize_cells(filename_report,'Alignment Summary',['F41','G41'],unbold=True,align='center',number_decimals=3)
        stylize_cells(filename_report,'Alignment Summary',['H41','H41'],unbold=True,align='center',thick_right=True,number_decimals=3)
    
    autofit_columns(filename_report,'Transformations')
    autofit_columns(filename_report,'USMN Raw')
    no_fill(filename_report,'Transformations')
    no_fill(filename_report,'USMN Raw')
    
    print("Stylizing report...")
    if module_type == 'F':
        remove_rows(filename_report,'Alignment Summary',row_bounds='31:33')
    elif module_type == 'Q':
        remove_rows(filename_report,'Alignment Summary',row_bounds='29:38')
    if num_magnets != 11:
        magnet_list_blank_rows = str(11+num_magnets) + ':21'
        remove_rows(filename_report,'Alignment Summary',row_bounds=magnet_list_blank_rows)
    
    remove_rows(filename_report,'Installation Fiducials',row_bounds='1:9')
    remove_rows(filename_report,'Transformations',row_bounds='1:9')
    remove_rows(filename_report,'USMN Raw',row_bounds='1:9')
    
    stylize_cells(filename_report,'Installation Fiducials',['A1','A1'],align='center',border=False)
    stylize_cells(filename_report,'Installation Fiducials',['C2','E3'],align='right',border=False)
    stylize_cells(filename_report,'Installation Fiducials',['A2','B100'],align='left',border=False)
    stylize_cells(filename_report,'Installation Fiducials',['C4','E100'],align='center',number_decimals=6,border=False)
    
    stylize_cells(filename_report,'Transformations',['A1','L700'], border=False)
    stylize_cells(filename_report,'USMN Raw',['A1','J450'], border=False)
    
    autosize_row_height(filename_report,'Installation Fiducials',size='small')
    autosize_row_height(filename_report,'Transformations')
    autosize_row_height(filename_report,'USMN Raw')
    wb = load_workbook(filename_report)
    wb.active = 0
    wb.save(filename_report)
    print("Assembly survey report created successfully...")

    savefile_to_pdf(filename_report)
    print("Alignment summary tab exported to PDF...")

    archive_filename = 'Archive\Report ' + module_name + ' Assembly Survey'
    os.system('copy \"' + module_name + '\Report ' + module_name + ' Assembly Survey.xlsx' + '\" \"' + archive_filename + '.xlsx' + '\"')
    os.system('copy \"' + module_name + '\Report ' + module_name + ' Assembly Survey.pdf' + '\" \"' + archive_filename + '.pdf' + '\"')
    print("Report saved to archive folder...")
    
    log_entry(filename_report,RMS_data)
    print("Entry created in log sheet...")
    print("Done!")

module_name = widgets.Text(value='', description='Enter module name:', disabled=False,
                            style = {'description_width': 'initial'}, layout=widgets.Layout(width="auto", height="auto"))
button = widgets.Button(description="Create assembly survey report", layout=widgets.Layout(width="auto", height="auto"))
button.on_click(on_button_clicked)